import csv
import io
import json
import logging
from typing import Any, Tuple

import pdfplumber
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _safe_decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def _rows_to_lines(rows: list[list[str]]) -> list[str]:
    lines = []
    for row in rows:
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        lines.append(" | ".join(cells))
    return lines


def _json_to_lines(obj: Any) -> list[str]:
    lines: list[str] = []

    def handle_entry(entry: dict):
        name = entry.get("name") or entry.get("artist") or entry.get("artist_name")
        platform = entry.get("platform") or entry.get("service")
        url = entry.get("platform_url") or entry.get("url") or entry.get("link")
        platform_id = entry.get("platform_id") or entry.get("channel_id")
        genres = entry.get("genre_tags") or entry.get("genres")

        parts = []
        if name:
            parts.append(str(name))
        if platform:
            parts.append(str(platform))
        if platform_id:
            parts.append(str(platform_id))
        if url:
            parts.append(str(url))
        if genres:
            if isinstance(genres, list):
                parts.append(", ".join(str(g) for g in genres))
            else:
                parts.append(str(genres))
        if parts:
            lines.append(" | ".join(parts))

    if isinstance(obj, dict):
        if isinstance(obj.get("artists"), list):
            for entry in obj["artists"]:
                if isinstance(entry, dict):
                    handle_entry(entry)
        elif isinstance(obj.get("roster"), list):
            for entry in obj["roster"]:
                if isinstance(entry, dict):
                    handle_entry(entry)
        else:
            # Fall back to stringifying the object
            lines.append(json.dumps(obj))
    elif isinstance(obj, list):
        for entry in obj:
            if isinstance(entry, dict):
                handle_entry(entry)
            else:
                lines.append(str(entry))
    else:
        lines.append(str(obj))

    return lines


def extract_text_from_upload(
    filename: str | None,
    content_type: str | None,
    data: bytes,
) -> Tuple[str, list[str]]:
    warnings: list[str] = []
    name = (filename or "").lower()

    if name.endswith(".csv") or (content_type == "text/csv"):
        text = _safe_decode(data)
        reader = csv.reader(io.StringIO(text))
        rows = [[str(c) for c in row] for row in reader]
        lines = _rows_to_lines(rows)
        return "\n".join(lines), warnings

    if name.endswith(".tsv") or (content_type == "text/tab-separated-values"):
        text = _safe_decode(data)
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        rows = [[str(c) for c in row] for row in reader]
        lines = _rows_to_lines(rows)
        return "\n".join(lines), warnings

    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" | ".join(cells))
        if not lines:
            warnings.append("No readable rows found in spreadsheet")
        return "\n".join(lines), warnings

    if name.endswith(".xls"):
        warnings.append("XLS is not supported; please export to XLSX or CSV")
        return "", warnings

    if name.endswith(".pdf") or (content_type == "application/pdf"):
        text_blocks: list[str] = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text:
                    text_blocks.append(page_text)
        if not text_blocks:
            warnings.append("No text found in PDF; consider exporting as text or CSV")
        return "\n".join(text_blocks), warnings

    if name.endswith(".json") or (content_type == "application/json"):
        text = _safe_decode(data)
        try:
            obj = json.loads(text)
        except Exception:
            warnings.append("JSON parse failed; using raw text")
            return text, warnings
        lines = _json_to_lines(obj)
        return "\n".join(lines), warnings

    # Fallback: treat as text
    text = _safe_decode(data)
    return text, warnings
